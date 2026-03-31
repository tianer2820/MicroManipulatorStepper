import csv
import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from dataclasses import dataclass, asdict, fields
from typer import Typer
from pathlib import Path
from datetime import datetime

APP = Typer()

DEFAULT_EXCEL_OUTPUT = Path("comparison_results.xlsx")
DEFAULT_MATCH_LIMIT = 0.8
DEFAULT_RANSAC_THRESHOLD = 5
MAX_WIDTH = 1600
MAX_HEIGHT = 900
DEFAULT_PIXEL_UM = 0.6209
DEFAULT_CAMERA_HEIGHT_UM = 1500.0   # height of camera above the plane (in µm)
DEFAULT_FOV_DEG = 60.0              # diagonal field of view in degrees


# ------------------------------------------------------------------ #
# Data classes for CSV output
# ------------------------------------------------------------------ #

@dataclass
class ImageComparisonResult:
    """Result from comparing two individual images."""
    image1: str
    image2: str
    tx_um: float
    ty_um: float
    tz_um: float
    yaw_deg: float
    total_movement_um: float
    matched_features: int = 0


@dataclass
class PairwiseComparisonResult:
    """Result from comparing all images across two locations."""
    location1_index: int | str
    location2_index: int | str
    location1_image: str
    location2_image: str
    tx_um: float
    ty_um: float
    tz_um: float
    yaw_deg: float
    total_movement_um: float


@dataclass
class AverageTransformResult:
    """Average transform statistics across multiple comparisons."""
    location1_index: int | str
    location2_index: int | str
    location1_image: str
    location2_image: str
    tx_um: float
    tx_std_um: float
    ty_um: float
    ty_std_um: float
    tz_um: float
    tz_std_um: float
    yaw_deg: float
    yaw_std_deg: float
    total_movement_um: float


@dataclass
class MovementResult:
    """Result from comparing two folder positions."""
    from_position: str
    to_position: str
    expected_distance_um: float
    actual_distance_um: float
    difference_um: float
    feed: float = 0.0  # Feed rate used for this movement


@dataclass
class MovementStatistics:
    """Statistics for movements of the same size and feed."""
    expected_distance_um: float
    feed: float
    count: int
    mean_actual_distance_um: float
    std_actual_distance_um: float
    mean_difference_um: float
    std_difference_um: float


@dataclass
class ImageVariationResult:
    """Result from comparing images at the same endpoint location."""
    endpoint: str  # Rounded endpoint coordinates
    num_visits: int  # Number of times this endpoint was visited
    num_image_pairs: int  # Total number of image pairs compared
    mean_tx_um: float  # Mean translation in X
    std_tx_um: float  # Std of translation in X
    mean_ty_um: float  # Mean translation in Y
    std_ty_um: float  # Std of translation in Y
    mean_total_xy_movement_um: float  # Mean total XY movement (ignoring Z and rotation)


def build_camera_intrinsics(img_shape: tuple, fov_deg: float) -> np.ndarray:
    """
    Build a simple pinhole camera intrinsic matrix from image size and diagonal FOV.

    Args:
        img_shape: (height, width) of the image in pixels.
        fov_deg:   Diagonal field-of-view in degrees.

    Returns:
        3×3 intrinsic matrix K.
    """
    h, w = img_shape[:2]
    diag_px = np.sqrt(w**2 + h**2)
    fov_rad = np.radians(fov_deg)
    # focal length from diagonal FOV
    f = (diag_px / 2.0) / np.tan(fov_rad / 2.0)
    cx, cy = w / 2.0, h / 2.0
    K = np.array([
        [f,  0, cx],
        [0,  f, cy],
        [0,  0,  1],
    ], dtype=np.float64)
    return K


def pixels_to_world(points_px: np.ndarray, K: np.ndarray, camera_height: float) -> np.ndarray:
    """
    Back-project 2-D image points onto the Z=0 ground plane, given the camera
    is at height `camera_height` above that plane (looking straight down).

    Returns an (N, 3) array of 3-D world points with Z=0.
    """
    K_inv = np.linalg.inv(K)
    pts_h = np.hstack([points_px, np.ones((len(points_px), 1))])   # (N,3) homogeneous
    rays = (K_inv @ pts_h.T).T                                       # (N,3) normalised ray dirs
    # intersect with Z=0 plane: world_point = camera_origin + t * ray
    # camera_origin = (0, 0, camera_height), ray direction has ray_z component
    # t = -camera_height / ray[:, 2]  (ray_z is negative for a downward-looking cam)
    t = camera_height / rays[:, 2]    # works when looking *toward* +Z ground plane
    world_pts = rays * t[:, np.newaxis]
    world_pts[:, 2] = 0.0             # enforce Z=0 on the ground plane
    return world_pts.astype(np.float64)


def compare_images(
    img1_path: Path,
    img2_path: Path,
    match_limit: float,
    camera_height: float,
    fov_deg: float,
    pixel_to_um: float,
):
    """
    Detect and match SIFT features between two images, then estimate the 3-D
    transform using a two-stage approach:

      Stage 1 — estimateAffinePartial2D (RANSAC):
        Recovers X/Y translation and in-plane rotation (yaw) accurately.
        The affine scale factor is used to derive Z.

      Stage 2 — Z from scale:
        A scale factor s != 1 means camera 2 was closer/farther than camera 1.
        Given the camera was at height H above the plane:
            new_height = H / s        (larger features → smaller height → negative ΔZ)
            ΔZ = new_height − H = H * (1/s − 1)

        solvePnP is NOT used for the full pose because its tvec.Z encodes the
        distance from the world plane origin, not the *change* in camera height,
        leading to large spurious Z values for pure lateral motion.
    """
    img1 = cv2.imread(str(img1_path), cv2.IMREAD_GRAYSCALE)
    img2 = cv2.imread(str(img2_path), cv2.IMREAD_GRAYSCALE)

    if img1 is None:
        raise FileNotFoundError(f"Could not read image: {img1_path}")
    if img2 is None:
        raise FileNotFoundError(f"Could not read image: {img2_path}")

    K1 = build_camera_intrinsics(img1.shape, fov_deg)
    K2 = build_camera_intrinsics(img2.shape, fov_deg)

    sift = cv2.SIFT_create(
        nfeatures=15000,
        contrastThreshold=0.01,
        edgeThreshold=5,
    )

    kp1, des1 = sift.detectAndCompute(img1, None)
    kp2, des2 = sift.detectAndCompute(img2, None)

    # FLANN matcher
    index_params = dict(algorithm=1, trees=5)
    search_params = dict(checks=50)
    flann = cv2.FlannBasedMatcher(index_params, search_params)
    raw_matches = flann.knnMatch(des1, des2, k=2)

    good = []
    for m, n in raw_matches:
        if m.distance < match_limit * n.distance:
            good.append(m)

    pts1_px = np.float32([kp1[m.queryIdx].pt for m in good])
    pts2_px = np.float32([kp2[m.trainIdx].pt for m in good])

    # ------------------------------------------------------------------ #
    # Stage 1: in-plane transform (X, Y, yaw) via partial affine
    # ------------------------------------------------------------------ #
    M, inlier_mask = cv2.estimateAffinePartial2D(
        pts1_px, pts2_px,
        method=cv2.RANSAC,
        ransacReprojThreshold=DEFAULT_RANSAC_THRESHOLD,
    )
    if M is None:
        raise RuntimeError("estimateAffinePartial2D failed — not enough matches.")

    inlier_matches = [m for m, keep in zip(good, inlier_mask.flatten()) if keep]
    inlier_matches = sorted(inlier_matches, key=lambda m: m.distance)

    # Decompose the 2×3 affine matrix
    # M = [[s·cosθ, -s·sinθ, tx],
    #      [s·sinθ,  s·cosθ, ty]]
    tx_px = M[0, 2]
    ty_px = M[1, 2]
    scale = np.sqrt(M[0, 0]**2 + M[1, 0]**2)   # uniform scale (should be ~1 for same-height)
    theta_rad = np.arctan2(M[1, 0], M[0, 0])

    # ------------------------------------------------------------------ #
    # Stage 2: Z from scale change
    #   scale > 1  → features appear larger in img2 → camera moved closer → ΔZ < 0
    #   scale < 1  → features appear smaller        → camera moved farther → ΔZ > 0
    #   ΔZ (in pixels) = camera_height_px * (1/scale − 1)
    # ------------------------------------------------------------------ #
    camera_height_px = camera_height / pixel_to_um
    dz_px = camera_height_px * (1.0 / scale - 1.0)

    # ------------------------------------------------------------------ #
    # Pack into a 4×4 homogeneous transform for downstream use
    # ------------------------------------------------------------------ #
    R_yaw = np.array([
        [ np.cos(theta_rad), -np.sin(theta_rad), 0],
        [ np.sin(theta_rad),  np.cos(theta_rad), 0],
        [0,                   0,                  1],
    ])
    T = np.eye(4)
    T[:3, :3] = R_yaw
    T[0,  3]  = tx_px
    T[1,  3]  = ty_px
    T[2,  3]  = dz_px

    return img1, img2, kp1, kp2, inlier_matches, T, K1, K2


def decompose_transform(T: np.ndarray, pixel_to_um: float):
    """
    Extract translation (X, Y, Z in µm) and in-plane rotation (yaw, degrees)
    from the 4x4 homogeneous transform produced by compare_images.

    T encodes:
      - T[0,3], T[1,3] : lateral translation in pixels  -> converted to um
      - T[2,3]         : Z shift in pixels               -> converted to um
                         (derived from affine scale: dZ = H*(1/s - 1))
      - T[:3,:3]       : pure yaw rotation matrix
    """
    tx_um = T[0, 3] * pixel_to_um
    ty_um = T[1, 3] * pixel_to_um
    tz_um = T[2, 3] * pixel_to_um

    # Yaw from the rotation matrix (rotation around Z axis only)
    yaw_deg = np.degrees(np.arctan2(T[1, 0], T[0, 0]))

    return (tx_um, ty_um, tz_um), yaw_deg


# ------------------------------------------------------------------ #
# Visualisation  (unchanged logic, kept for convenience)
# ------------------------------------------------------------------ #

def visualize_matches(img1, img2, kp1, kp2, matches):
    img1_color = cv2.cvtColor(img1, cv2.COLOR_GRAY2BGR)
    img2_color = cv2.cvtColor(img2, cv2.COLOR_GRAY2BGR)

    h1, w1 = img1_color.shape[:2]
    h2, w2 = img2_color.shape[:2]

    max_h = max(h1, h2)
    max_w = max(w1, w2)

    if h1 < max_h:
        img1_color = cv2.copyMakeBorder(img1_color, 0, max_h - h1, 0, 0, cv2.BORDER_CONSTANT)
    if w1 < max_w:
        img1_color = cv2.copyMakeBorder(img1_color, 0, 0, max_w - w1, 0, cv2.BORDER_CONSTANT)
    if h2 < max_h:
        img2_color = cv2.copyMakeBorder(img2_color, 0, max_h - h2, 0, 0, cv2.BORDER_CONSTANT)
    if w2 < max_w:
        img2_color = cv2.copyMakeBorder(img2_color, 0, 0, max_w - w2, 0, cv2.BORDER_CONSTANT)

    combined = np.hstack((img1_color, img2_color))
    h, w = combined.shape[:2]
    scale = min(MAX_WIDTH / w, MAX_HEIGHT / h, 1.0)
    if scale < 1.0:
        combined = cv2.resize(combined, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_AREA)

    base = combined.copy()
    show_matches = True
    match_to_color = {}

    window_name = "Feature Matches — 3D (m=toggle, q=quit)"
    cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)

    while True:
        if cv2.getWindowProperty(window_name, cv2.WND_PROP_VISIBLE) < 1:
            break

        display = base.copy()

        if show_matches:
            left_img_width = display.shape[1] / 2
            for m in matches[:100]:
                x1, y1 = kp1[m.queryIdx].pt
                x2, y2 = kp2[m.trainIdx].pt

                dh, dw = display.shape[:2]
                y1s, x1s = dh / h1, (dw / 2) / w1
                y2s, x2s = dh / h2, (dw / 2) / w2

                pt1 = (int(x1 * x1s), int(y1 * y1s))
                pt2_shifted = (int(x2 * x2s + left_img_width), int(y2 * y2s))

                if m.queryIdx not in match_to_color:
                    match_to_color[m.queryIdx] = tuple(np.random.randint(0, 255, 3).tolist())
                color = match_to_color[m.queryIdx]

                cv2.circle(display, pt1, 4, color, -1)
                cv2.circle(display, pt2_shifted, 4, color, -1)
                cv2.line(display, pt1, pt2_shifted, color, 1)

        cv2.imshow(window_name, display)
        key = cv2.waitKey(1) & 0xFF

        if key == ord('m'):
            show_matches = not show_matches
        elif key == ord('q') or key == 27:
            break

    cv2.destroyAllWindows()


# ------------------------------------------------------------------ #
# CLI entry point
# ------------------------------------------------------------------ #

def compute_average_transform(
    image_paths: list,
    match_limit: float,
    camera_height: float,
    fov_deg: float,
    pixel_to_um: float,
) -> np.ndarray:
    """
    Compute the average transform across multiple consecutive images at the same location.
    
    Each image is compared to the next, and all transforms are averaged to get a
    representative pose for that location set.
    
    Args:
        image_paths: List of Path objects for images at the same location
        match_limit: SIFT feature match threshold
        camera_height: Camera height in µm
        fov_deg: Field of view in degrees
        pixel_to_um: Pixel to micrometer conversion factor
        
    Returns:
        Average 4×4 homogeneous transform matrix
    """
    if len(image_paths) < 2:
        raise ValueError("Need at least 2 images to compute average transform")
    
    transforms = []
    
    # Compare each consecutive pair of images
    for i in range(len(image_paths)):
        for j in range(len(image_paths)):
            if j == i:
                continue
            # Skip j,i and i,j duplicates by enforcing j > i
            if j < i:
                continue

            try:
                _, _, _, _, _, T, _, _ = compare_images(
                    image_paths[i],
                    image_paths[j],
                    match_limit,
                    camera_height,
                    fov_deg,
                    pixel_to_um,
                )
                transforms.append(T)
                print(f"  Compared image {i} to {j}: OK")
            except Exception as e:
                print(f"  Warning: Failed to compare image {i} to {j}: {e}")
                continue
    
    if not transforms:
        raise RuntimeError("No valid transforms computed from image set")
    
    # Average the translation components
    avg_translation = np.mean([T[:3, 3] for T in transforms], axis=0)
    std_translation = np.std([T[:3, 3] for T in transforms], axis=0)
    
    # Average the rotation matrices using matrix mean (Frobenius norm)
    avg_rotation = np.mean([T[:3, :3] for T in transforms], axis=0)
    U, _, Vt = np.linalg.svd(avg_rotation)
    avg_rotation = U @ Vt
    
    # Build the average transform
    avg_T = np.eye(4)
    avg_T[:3, :3] = avg_rotation
    avg_T[:3, 3] = avg_translation
    
    # Decompose all transforms to get stats on individual components
    decomposed = [decompose_transform(T, pixel_to_um) for T in transforms]
    translations = np.array([(t[0], t[1], t[2]) for t, _ in decomposed])
    yaws = np.array([y for _, y in decomposed])
    
    # Print debugging info for average transform
    (tx_avg, ty_avg, tz_avg), yaw_avg = decompose_transform(avg_T, pixel_to_um)
    print(f"\n  Average transform across {len(transforms)} comparisons:")
    print(f"    Translation: X = {tx_avg:+.2f} ± {np.std(translations[:, 0]):.2f} um,  Y = {ty_avg:+.2f} ± {np.std(translations[:, 1]):.2f} um,  Z = {tz_avg:+.2f} ± {np.std(translations[:, 2]):.2f} um")
    print(f"    Total movement : {np.sqrt(tx_avg**2 + ty_avg**2):.2f} um")
    print(f"    Rotation:    yaw = {yaw_avg:+.2f} ± {np.std(yaws):.2f} deg")
    
    stats = {
        'tx_mean': tx_avg,
        'tx_std': np.std(translations[:, 0]),
        'ty_mean': ty_avg,
        'ty_std': np.std(translations[:, 1]),
        'tz_mean': tz_avg,
        'tz_std': np.std(translations[:, 2]),
        'yaw_mean': yaw_avg,
        'yaw_std': np.std(yaws),
        'total_movement_mean': np.sqrt(tx_avg**2 + ty_avg**2),
    }
    
    return avg_T, stats


def write_results_to_excel(
    filepath: Path,
    location_averages: dict = None,
    pairwise_results: list = None,
    summary_result: AverageTransformResult = None,
):
    """
    Write transform results to an Excel file with multiple sheets.
    Appends to existing file or creates new one if it doesn't exist.
    
    Args:
        filepath: Path to the Excel file to write
        location_averages: Dict with keys 'location1' and 'location2' containing stats dicts
        pairwise_results: List of PairwiseComparisonResult dataclass instances
        summary_result: AverageTransformResult dataclass instance for summary
    """
    filepath = Path(filepath)
    
    # Load existing workbook or create new one
    if filepath.exists():
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
    
    # Create or get Location Averages sheet
    if location_averages:
        sheet_name = "Location Averages"
        if sheet_name in wb.sheetnames:
            ws_avg = wb[sheet_name]
        else:
            ws_avg = wb.create_sheet(sheet_name)
            headers = ['Location', 'TX (µm)', 'TX Std (µm)', 'TY (µm)', 'TY Std (µm)', 
                       'TZ (µm)', 'TZ Std (µm)', 'Yaw (deg)', 'Yaw Std (deg)', 'Total Movement (µm)']
            ws_avg.append(headers)

        
        names_to_skip = set()
        for row in ws_avg.iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None:
                names_to_skip.add(row[0])

        # Add location data
        for location_name, stats in location_averages.items():
            if location_name in names_to_skip:
                print(f"  Skipping existing entry for {location_name} in Location Averages sheet")
                continue
            ws_avg.append([
                location_name,
                stats.get('tx_mean', ''),
                stats.get('tx_std', ''),
                stats.get('ty_mean', ''),
                stats.get('ty_std', ''),
                stats.get('tz_mean', ''),
                stats.get('tz_std', ''),
                stats.get('yaw_mean', ''),
                stats.get('yaw_std', ''),
                stats.get('total_movement_mean', ''),
            ])
    
    # Create or append to Pairwise Comparisons sheet
    if pairwise_results:
        sheet_name = "Pairwise Comparisons"
        if sheet_name in wb.sheetnames:
            ws_pair = wb[sheet_name]
            start_row = ws_pair.max_row + 1
        else:
            ws_pair = wb.create_sheet(sheet_name)
            headers = ['Location1 Index', 'Location2 Index', 'Location1 Image', 'Location2 Image',
                       'TX (µm)', 'TY (µm)', 'TZ (µm)', 'Yaw (deg)', 'Total Movement (µm)']
            ws_pair.append(headers)
            start_row = 2
        
        names_to_maybe_skip = set()
        for row in ws_pair.iter_rows(min_row=2, values_only=True):
            if row and row[2] is not None and row[3] is not None:
                names_to_maybe_skip.add((row[2], row[3]))

        # Add pairwise comparison data
        for result in pairwise_results:
            if (result.location1_image, result.location2_image) in names_to_maybe_skip:
                print(f"  Skipping existing entry for {result.location1_image} vs {result.location2_image} in Pairwise Comparisons sheet")
                continue
            ws_pair.append([
                result.location1_index,
                result.location2_index,
                result.location1_image,
                result.location2_image,
                result.tx_um,
                result.ty_um,
                result.tz_um,
                result.yaw_deg,
                result.total_movement_um,
            ])
    
    # Create or update Summary sheet
    if summary_result:
        sheet_name = "Summary"
        if sheet_name in wb.sheetnames:
            ws_summary = wb[sheet_name]
            start_row = ws_summary.max_row + 1
        else:
            ws_summary = wb.create_sheet(sheet_name)
            headers = ["Location1", "Location2", "TX (µm)", "TX Std (µm)", "TY (µm)", "TY Std (µm)",
                       "TZ (µm)", "TZ Std (µm)", "Yaw (deg)", "Yaw Std (deg)", "Total Movement (µm)"]
            ws_summary.append(headers)
            start_row = 2


        ws_summary.append([
            summary_result.location1_image,
            summary_result.location2_image,
            summary_result.tx_um,
            summary_result.tx_std_um,
            summary_result.ty_um,
            summary_result.ty_std_um,
            summary_result.tz_um,
            summary_result.tz_std_um,
            summary_result.yaw_deg,
            summary_result.yaw_std_deg,
            summary_result.total_movement_um,
        ])
    
    wb.save(filepath)
    print(f"  Results written to: {filepath}")


@APP.command()
def compare_images_script(
    img1_path: Path,
    img2_path: Path,
    visualize: bool = True,
    match_limit: float = DEFAULT_MATCH_LIMIT,
    pixel_to_um: float = DEFAULT_PIXEL_UM,
    camera_height_um: float = DEFAULT_CAMERA_HEIGHT_UM,
    fov_deg: float = DEFAULT_FOV_DEG,
    excel_output: Path = None,
):
    img1, img2, kp1, kp2, matches, T, K1, K2 = compare_images(
        img1_path, img2_path, match_limit, camera_height_um, fov_deg, pixel_to_um
    )

    (tx, ty, tz), yaw_deg = decompose_transform(T, pixel_to_um)

    print(f"Matched features (inliers): {len(matches)}")
    print(f"Translation : X = {tx:+.2f} um,  Y = {ty:+.2f} um,  Z = {tz:+.2f} um")
    print(f"Total movement : {np.sqrt(tx**2 + ty**2):.2f} um")
    print(f"Rotation    : yaw = {yaw_deg:+.2f} deg")

    if visualize:
        visualize_matches(img1, img2, kp1, kp2, matches)


@APP.command()
def compare_locations(
    location1_dir: Path,
    location2_dir: Path,
    visualize: bool = True,
    match_limit: float = DEFAULT_MATCH_LIMIT,
    pixel_to_um: float = DEFAULT_PIXEL_UM,
    camera_height_um: float = DEFAULT_CAMERA_HEIGHT_UM,
    fov_deg: float = DEFAULT_FOV_DEG,
    excel_output: Path | None = None,
):
    """
    Compare two locations by computing the average pose at each location,
    then finding the transform between them.
    
    Args:
        location1_dir: Directory containing images from location 1
        location2_dir: Directory containing images from location 2
        visualize: Whether to show feature matches between representative images
        match_limit: SIFT feature match threshold (0.0-1.0)
        pixel_to_um: Pixel to micrometer conversion factor
        camera_height_um: Camera height in µm
        fov_deg: Field of view in degrees
        excel_output: Optional path to Excel file for saving results
    """
    # Collect image files from each location
    image_extensions = {'.png', '.jpg', '.jpeg', '.bmp', '.tiff'}
    
    location1_images = sorted([
        p for p in location1_dir.glob('*')
        if p.suffix.lower() in image_extensions
    ])
    location2_images = sorted([
        p for p in location2_dir.glob('*')
        if p.suffix.lower() in image_extensions
    ])
    
    if not location1_images:
        raise FileNotFoundError(f"No images found in {location1_dir}")
    if not location2_images:
        raise FileNotFoundError(f"No images found in {location2_dir}")
    
    print(f"\nLocation 1: Found {len(location1_images)} images")
    print(f"Location 2: Found {len(location2_images)} images")
    
    # Compute average pose for each location
    print("\nComputing average pose for Location 1...")
    avg_T1, stats1 = compute_average_transform(
        location1_images,
        match_limit,
        camera_height_um,
        fov_deg,
        pixel_to_um,
    )
    
    print("\nComputing average pose for Location 2...")
    avg_T2, stats2 = compute_average_transform(
        location2_images,
        match_limit,
        camera_height_um,
        fov_deg,
        pixel_to_um,
    )
    
    
    # Compare all images from location 1 with all images from location 2
    print("\n" + "="*60)
    print("COMPARING ALL IMAGES ACROSS LOCATIONS")
    print("="*60)
    
    all_transforms = []
    comparison_results = []
    comparison_count = 0
    successful_comparisons = 0
    
    for i, img1_path in enumerate(location1_images):
        for j, img2_path in enumerate(location2_images):
            comparison_count += 1
            try:
                _, _, _, _, _, T, _, _ = compare_images(
                    img1_path,
                    img2_path,
                    match_limit,
                    camera_height_um,
                    fov_deg,
                    pixel_to_um,
                )
                all_transforms.append(T)
                successful_comparisons += 1
                (tx_pair, ty_pair, tz_pair), yaw_pair = decompose_transform(T, pixel_to_um)
                print(f"  [{i}→{j}] ΔX = {tx_pair:+7.1f} um, ΔY = {ty_pair:+7.1f} um, ΔZ = {tz_pair:+7.1f} um, Δyaw = {yaw_pair:+6.1f}° total = {np.sqrt(tx_pair**2 + ty_pair**2):.1f} um")
                
                # Collect result for CSV
                comparison_results.append(
                    PairwiseComparisonResult(
                        location1_index=i,
                        location2_index=j,
                        location1_image=str(img1_path.name),
                        location2_image=str(img2_path.name),
                        tx_um=tx_pair,
                        ty_um=ty_pair,
                        tz_um=tz_pair,
                        yaw_deg=yaw_pair,
                        total_movement_um=np.sqrt(tx_pair**2 + ty_pair**2),
                    )
                )
            except Exception as e:
                print(f"  [{i}→{j}] Failed: {e}")
                continue
    
    print(f"\nCompleted {successful_comparisons}/{comparison_count} comparisons")
    
    if all_transforms:
        # Compute average across all pairwise comparisons
        avg_translation_all = np.mean([T[:3, 3] for T in all_transforms], axis=0)
        std_translation_all = np.std([T[:3, 3] for T in all_transforms], axis=0)
        avg_rotation_all = np.mean([T[:3, :3] for T in all_transforms], axis=0)
        U, _, Vt = np.linalg.svd(avg_rotation_all)
        avg_rotation_all = U @ Vt
        
        avg_T_all = np.eye(4)
        avg_T_all[:3, :3] = avg_rotation_all
        avg_T_all[:3, 3] = avg_translation_all
        
        # Decompose all to get stats
        decomposed_all = [decompose_transform(T, pixel_to_um) for T in all_transforms]
        translations_all = np.array([(t[0], t[1], t[2]) for t, _ in decomposed_all])
        yaws_all = np.array([y for _, y in decomposed_all])
        
        (tx_all, ty_all, tz_all), yaw_all = decompose_transform(avg_T_all, pixel_to_um)
        
        print("\n" + "="*60)
        print("AVERAGE TRANSFORM FROM COMPARISONS")
        print("="*60)
        print(f"Translation : X = {tx_all:+.2f} ± {np.std(translations_all[:, 0]):.2f} um,  Y = {ty_all:+.2f} ± {np.std(translations_all[:, 1]):.2f} um,  Z = {tz_all:+.2f} ± {np.std(translations_all[:, 2]):.2f} um")
        print(f"Total movement : {np.sqrt(tx_all**2 + ty_all**2):.2f} um")
        print(f"Rotation    : yaw = {yaw_all:+.2f} ± {np.std(yaws_all):.2f} deg")
        print("="*60)
        
        # Create summary result
        summary_result = AverageTransformResult(
            location1_index='AVERAGE',
            location2_index='AVERAGE',
            location1_image=location1_dir.name,
            location2_image=location2_dir.name,
            tx_um=tx_all,
            tx_std_um=np.std(translations_all[:, 0]),
            ty_um=ty_all,
            ty_std_um=np.std(translations_all[:, 1]),
            tz_um=tz_all,
            tz_std_um=np.std(translations_all[:, 2]),
            yaw_deg=yaw_all,
            yaw_std_deg=np.std(yaws_all),
            total_movement_um=np.sqrt(tx_all**2 + ty_all**2),
        )
        
        # Write Excel output if requested
        if excel_output:
            print(f"\nWriting results to Excel...")
            location_averages = {
                location1_dir.name: stats1,
                location2_dir.name: stats2,
            }
            write_results_to_excel(
                excel_output,
                location_averages=location_averages,
                pairwise_results=comparison_results,
                summary_result=summary_result,
            )
    
    # Optionally visualize matches between first images of each location
    if visualize and location1_images and location2_images:
        print("\nGenerating visualization of matches between first images...")
        try:
            img1, img2, kp1, kp2, matches, _, _, _ = compare_images(
                location1_images[0],
                location2_images[0],
                match_limit,
                camera_height_um,
                fov_deg,
                pixel_to_um,
            )
            visualize_matches(img1, img2, kp1, kp2, matches)
        except Exception as e:
            print(f"Could not visualize matches: {e}")

@dataclass
class FolderInfo:
    idx: int
    expected: float
    startx: float
    starty: float
    endx: float
    endy: float


ENDPOINT_ROUNDING_MM = 0.00001  # 10 nanometers in mm


def round_endpoint(x: float, y: float, tolerance_mm: float = ENDPOINT_ROUNDING_MM) -> tuple[float, float]:
    """
    Round endpoint coordinates to the nearest tolerance value.
    This helps group visits to essentially the same location.
    
    Args:
        x, y: Coordinates in mm
        tolerance_mm: Rounding tolerance (default: 10 nm = 0.00001 mm)
    
    Returns:
        Tuple of rounded (x, y) coordinates
    """
    rounded_x = round(x / tolerance_mm) * tolerance_mm
    rounded_y = round(y / tolerance_mm) * tolerance_mm
    return rounded_x, rounded_y


def parse_position(folder_name: str) -> FolderInfo:
    """
    Parse folder name of format 'idx_expected_startx_starty_endx_endy' into coordinates.
    Returns a FolderInfo object with all parsed values, handling decimal notation like '0+1' -> 0.1
    """
    parts = folder_name.split('_')
    if len(parts) != 6:
        raise ValueError(f"Invalid folder format: {folder_name}. Expected 'idx_expected_startx_starty_endx_endy'")
    
    try:
        idx = int(parts[0])
        # Convert notation like '0+1' to '0.1'
        expected = float(parts[1].replace('d', '.'))
        startx = float(parts[2].replace('d', '.'))
        starty = float(parts[3].replace('d', '.'))
        endx = float(parts[4].replace('d', '.'))
        endy = float(parts[5].replace('d', '.'))
        return FolderInfo(idx, expected, startx, starty, endx, endy)
    except (ValueError, IndexError) as e:
        raise ValueError(f"Could not parse folder name {folder_name}: {e}")


def calculate_xy_distance(pos1: FolderInfo, pos2: FolderInfo) -> float:
    """
    Calculate 2D distance between two positions in micrometers.
    Positions are (idx, startx, starty, endx, endy) tuples in mm, returns distance in µm.
    Uses the end coordinates of pos1 compared to start coordinates of pos2.
    """
    x1 = pos2[1]  # endx
    y1 = pos2[2]  # endy
    x2 = pos1[1]  # startx
    y2 = pos1[2]  # starty
    
    dx = (x2 - x1) *1000
    dy = (y2 - y1) *1000
    return np.sqrt(dx**2 + dy**2)


@APP.command()
def analyze_movements(
    root_dir: Path,
    excel_output: Path = DEFAULT_EXCEL_OUTPUT,
    pixel_to_um: float = DEFAULT_PIXEL_UM,
    camera_height_um: float = DEFAULT_CAMERA_HEIGHT_UM,
    fov_deg: float = DEFAULT_FOV_DEG,
    match_limit: float = DEFAULT_MATCH_LIMIT,
):
    """
    Analyze movement accuracy by reading indexed image folders and CSV metadata.
    
    Folder structure: indexed folders (0, 1, 2, ...) containing images.
    Reads from:
    - idx_to_coord.csv: Maps index to (x, y) coordinates
    - path_order.csv: Defines movements (start_idx, end_idx, expected_distance_um)
    
    Args:
        root_dir: Root directory containing indexed folders and CSV files
        excel_output: Optional path to Excel file for saving results
        pixel_to_um: Pixel to micrometer conversion factor
        camera_height_um: Camera height in µm
        fov_deg: Field of view in degrees
        match_limit: SIFT feature match threshold
    """
    root_dir = Path(root_dir)
    
    # Read the CSV files
    idx_to_coord_file = root_dir / "idx_to_coord.csv"
    path_order_file = root_dir / "path_order.csv"
    
    if not idx_to_coord_file.exists():
        raise ValueError(f"File not found: {idx_to_coord_file}")
    if not path_order_file.exists():
        raise ValueError(f"File not found: {path_order_file}")
    
    # Load idx_to_coord mapping
    idx_to_coord = {}
    with open(idx_to_coord_file, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            idx = int(row['idx'])
            x = float(row['x'])
            y = float(row['y'])
            idx_to_coord[idx] = (x, y)
    
    print(f"\nLoaded {len(idx_to_coord)} location coordinates")
    for idx in sorted(idx_to_coord.keys()):
        x, y = idx_to_coord[idx]
        print(f"  Location {idx}: ({x:.6f}, {y:.6f}) mm")
    
    # Load path_order movements
    movements = []
    with open(path_order_file, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            start_idx = int(row['start_idx'])
            end_idx = int(row['end_idx'])
            expected_distance_mm = float(row['expected_distance']) if 'expected_distance' in row else None
            feed = float(row['feed']) if 'feed' in row else None
            movements.append((start_idx, end_idx, expected_distance_mm, feed))
    
    print(f"\nLoaded {len(movements)} movements")
    for start_idx, end_idx, expected_dist, feed in movements:
        if expected_dist is not None:
            print(f"  {start_idx} → {end_idx}: expected {expected_dist:.6f} mm ({expected_dist*1000:.1f} µm), feed={feed}")
        else:
            print(f"  {start_idx} → {end_idx}: no expected distance, feed={feed}")
    
    # Compare sequential positions
    movement_results: list[MovementResult] = []
    print("\n" + "="*80)
    print("COMPARING SEQUENTIAL POSITIONS")
    print("="*80)
    
    image_extensions = {'.png', '.jpg', '.jpeg', '.bmp', '.tiff'}
    
    for start_idx, end_idx, expected_distance_mm, feed in movements:
        # Convert expected distance from mm to µm
        expected_distance_um = expected_distance_mm * 1000 if expected_distance_mm is not None else None
        
        folder1 = root_dir / str(start_idx)
        folder2 = root_dir / str(end_idx)
        
        if not folder1.exists() or not folder2.exists():
            print(f"  [{start_idx} → {end_idx}] Skipped: folder(s) not found")
            continue
        
        # Get all images from each folder
        images1 = sorted([p for p in folder1.glob('*') if p.suffix.lower() in image_extensions])
        images2 = sorted([p for p in folder2.glob('*') if p.suffix.lower() in image_extensions])
        
        if not images1 or not images2:
            print(f"  [{start_idx} → {end_idx}] Skipped: missing images")
            continue
        
        try:
            # Compare all images from folder1 to all images from folder2
            actual_distances = []
            failed_comparisons = 0
            
            for img1_path in images1:
                for img2_path in images2:
                    try:
                        img1, img2, kp1, kp2, matches, T, K1, K2 = compare_images(
                            img1_path, img2_path, match_limit, camera_height_um, fov_deg, pixel_to_um
                        )
                        (tx, ty, tz), yaw = decompose_transform(T, pixel_to_um)
                        # Calculate actual distance (XY only)
                        distance = np.sqrt(tx**2 + ty**2)
                        actual_distances.append(distance)
                    except Exception as e:
                        failed_comparisons += 1
                        continue
            
            if not actual_distances:
                print(f"  [{start_idx} → {end_idx}] Failed: no successful image comparisons")
                continue
            
            # Use average of all comparisons
            actual_distance = np.mean(actual_distances)
            
            # Get coordinates
            x1, y1 = idx_to_coord[start_idx]
            x2, y2 = idx_to_coord[end_idx]
            
            # Calculate difference if expected distance is available
            if expected_distance_um is None:
                expected_distance_um = np.sqrt((x2 - x1)**2 + (y2 - y1)**2)
            difference = actual_distance - expected_distance_um

            result = MovementResult(
                from_position=f"({x1:.9f},{y1:.9f})",
                to_position=f"({x2:.9f},{y2:.9f})",
                expected_distance_um=expected_distance_um,
                actual_distance_um=actual_distance,
                difference_um=difference,
                feed=feed if feed is not None else 0.0,
            )
            movement_results.append(result)
            
            print(f"  [{start_idx} → {end_idx}]")
            print(f"    Expected: {expected_distance_um:7.1f} µm")
            print(f"    Actual:   {actual_distance:7.1f} µm")
            if expected_distance_um > 0:
                print(f"    Diff:     {difference:+7.1f} µm ({difference/expected_distance_um*100:+.1f}%)")
            
        except Exception as e:
            print(f"  [{start_idx} → {end_idx}] Failed: {e}")
            continue
    
    if not movement_results:
        print("No valid movement comparisons found!")
        return
    
    # Calculate statistics by movement size
    print("\n" + "="*80)
    print("STATISTICS BY MOVEMENT SIZE")
    print("="*80)
    
    # Group by (distance, feed) combination
    movement_by_size_and_feed: dict[tuple, list[MovementResult]] = {}
    for result in movement_results:
        size = round(result.expected_distance_um, 0)  # Round to nearest µm
        feed_key = round(result.feed, 1)  # Round feed to nearest 0.1
        key = (size, feed_key)
        if key not in movement_by_size_and_feed:
            movement_by_size_and_feed[key] = []
        movement_by_size_and_feed[key].append(result)
    
    statistics = []
    for (size, feed_key) in sorted(movement_by_size_and_feed.keys()):
        results = movement_by_size_and_feed[(size, feed_key)]
        actual_distances = [r.actual_distance_um for r in results]
        differences = [r.difference_um for r in results]
        
        stats = MovementStatistics(
            expected_distance_um=size,
            feed=feed_key,
            count=len(results),
            mean_actual_distance_um=np.mean(actual_distances),
            std_actual_distance_um=np.std(actual_distances),
            mean_difference_um=np.mean(differences),
            std_difference_um=np.std(differences),
        )
        statistics.append(stats)
        
        print(f"\n{size:6.0f} µm moves at feed {feed_key:7.1f} (n={len(results)}):")
        print(f"  Actual:     {np.mean(actual_distances):7.1f} ± {np.std(actual_distances):6.1f} µm")
        print(f"  Difference: {np.mean(differences):+7.1f} ± {np.std(differences):6.1f} µm")
    
    # Analyze image variations at same endpoints
    print("\n" + "="*80)
    print("ANALYZING IMAGE VARIATIONS AT SAME ENDPOINTS")
    print("="*80)
    
    image_variations: list[ImageVariationResult] = []
    endpoints_to_indices: dict[tuple, list[int]] = {}
    
    # Group indices by their rounded endpoint
    for idx, (x, y) in idx_to_coord.items():
        rounded_ep = round_endpoint(x, y)
        if rounded_ep not in endpoints_to_indices:
            endpoints_to_indices[rounded_ep] = []
        endpoints_to_indices[rounded_ep].append(idx)
    
    # Analyze variations for endpoints with multiple visits
    for endpoint, idx_list in sorted(endpoints_to_indices.items()):
        if len(idx_list) < 2:
            continue  # Skip endpoints with only one visit
        
        rounded_x, rounded_y = endpoint
        endpoint_str = f"({rounded_x:.9f},{rounded_y:.9f})"
        
        print(f"\n  Endpoint {endpoint_str}: {len(idx_list)} visits (indices: {idx_list})")
        
        # Collect all images from all visits to this endpoint
        all_images_by_idx: dict[int, list[Path]] = {}
        
        for idx in idx_list:
            folder = root_dir / str(idx)
            if folder.exists():
                images = sorted([p for p in folder.glob('*') if p.suffix.lower() in image_extensions])
                if images:
                    all_images_by_idx[idx] = images
        
        # Compare images across different visits
        all_comparisons = []
        idx_list_with_images = list(all_images_by_idx.keys())
        
        for i in range(len(idx_list_with_images)):
            for j in range(i + 1, len(idx_list_with_images)):
                idx_i = idx_list_with_images[i]
                idx_j = idx_list_with_images[j]
                
                images_i = all_images_by_idx[idx_i]
                images_j = all_images_by_idx[idx_j]
                
                # Compare all image pairs between these two visits
                for img_i in images_i:
                    for img_j in images_j:
                        try:
                            img1, img2, kp1, kp2, matches, T, K1, K2 = compare_images(
                                img_i, img_j, match_limit, camera_height_um, fov_deg, pixel_to_um
                            )
                            (tx, ty, tz), yaw = decompose_transform(T, pixel_to_um)
                            xy_movement = np.sqrt(tx**2 + ty**2)  # Only XY movement
                            
                            all_comparisons.append({
                                'tx': tx,
                                'ty': ty,
                                'xy_movement': xy_movement,
                            })
                        except Exception as e:
                            continue
        
        if all_comparisons:
            # Calculate statistics (XY only, ignoring Z and rotation)
            tx_values = [c['tx'] for c in all_comparisons]
            ty_values = [c['ty'] for c in all_comparisons]
            xy_movement_values = [c['xy_movement'] for c in all_comparisons]
            
            result = ImageVariationResult(
                endpoint=endpoint_str,
                num_visits=len(idx_list),
                num_image_pairs=len(all_comparisons),
                mean_tx_um=float(np.mean(tx_values)),
                std_tx_um=float(np.std(tx_values)),
                mean_ty_um=float(np.mean(ty_values)),
                std_ty_um=float(np.std(ty_values)),
                mean_total_xy_movement_um=float(np.mean(xy_movement_values)),
            )
            image_variations.append(result)
            
            print(f"    Compared {len(all_comparisons)} image pairs")
            print(f"    TX variation: {np.mean(tx_values):+7.2f} ± {np.std(tx_values):6.2f} µm")
            print(f"    TY variation: {np.mean(ty_values):+7.2f} ± {np.std(ty_values):6.2f} µm")
            print(f"    XY movement: {np.mean(xy_movement_values):7.2f} ± {np.std(xy_movement_values):6.2f} µm")
    
    # Write Excel output if requested
    if excel_output:
        print(f"\nWriting results to Excel...")
        write_movement_results_to_excel(excel_output, movement_results, statistics, image_variations)


def write_movement_results_to_excel(filepath: Path, movement_results: list, statistics: list, image_variations: list = None):
    """
    Write movement analysis results to Excel with three sheets.
    
    Sheet 1: Individual movements
    Sheet 2: Statistics by movement size
    Sheet 3: Image variations at same endpoints (if available)
    """
    filepath = Path(filepath)
    
    # Load existing workbook or create new one
    if filepath.exists():
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
        wb.remove(wb.active)
    
    # Sheet 1: Individual Movements
    sheet_name = "Movements"
    if sheet_name in wb.sheetnames:
        ws_moves = wb[sheet_name]
        # Clear existing data
        for row in ws_moves.iter_rows(min_row=2, max_row=ws_moves.max_row):
            for cell in row:
                cell.value = None
    else:
        ws_moves = wb.create_sheet(sheet_name)
        headers = ['From Position', 'To Position', 'Expected (µm)', 'Actual (µm)', 'Difference (µm)', 'Feed']
        ws_moves.append(headers)
    
    for result in movement_results:
        ws_moves.append([
            result.from_position,
            result.to_position,
            result.expected_distance_um,
            result.actual_distance_um,
            result.difference_um,
            result.feed,
        ])
    
    # Sheet 2: Statistics by Movement Size and Feed
    sheet_name = "Movement Statistics"
    if sheet_name in wb.sheetnames:
        ws_stats = wb[sheet_name]
        # Clear existing data
        for row in ws_stats.iter_rows(min_row=2, max_row=ws_stats.max_row):
            for cell in row:
                cell.value = None
    else:
        ws_stats = wb.create_sheet(sheet_name)
        headers = ['Expected (µm)', 'Feed', 'Count', 'Mean Actual (µm)', 'Std Actual (µm)', 
                   'Mean Difference (µm)', 'Std Difference (µm)']
        ws_stats.append(headers)
    
    for stat in statistics:
        ws_stats.append([
            stat.expected_distance_um,
            stat.feed,
            stat.count,
            stat.mean_actual_distance_um,
            stat.std_actual_distance_um,
            stat.mean_difference_um,
            stat.std_difference_um,
        ])
    
    # Sheet 3: Image Variations at Same Endpoints
    if image_variations:
        sheet_name = "Image Variations"
        if sheet_name in wb.sheetnames:
            ws_variations = wb[sheet_name]
            # Clear existing data
            for row in ws_variations.iter_rows(min_row=2, max_row=ws_variations.max_row):
                for cell in row:
                    cell.value = None
        else:
            ws_variations = wb.create_sheet(sheet_name)
            headers = [
                'Endpoint', 'Visits', 'Image Pairs',
                'Mean TX (µm)', 'Std TX (µm)',
                'Mean TY (µm)', 'Std TY (µm)',
                'Mean XY Movement (µm)'
            ]
            ws_variations.append(headers)
        
        for variation in image_variations:
            ws_variations.append([
                variation.endpoint,
                variation.num_visits,
                variation.num_image_pairs,
                variation.mean_tx_um,
                variation.std_tx_um,
                variation.mean_ty_um,
                variation.std_ty_um,
                variation.mean_total_xy_movement_um,
            ])
    
    wb.save(filepath)
    print(f"  Results written to: {filepath}")


if __name__ == "__main__":
    APP()